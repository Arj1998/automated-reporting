"""
report_generator.py
Automated MIS Report Generator: PostgreSQL → styled Excel

Author : Arijit Pani
GitHub : https://github.com/Arj1998

Run:
    python report_generator.py
"""

import os
import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# ── Config ────────────────────────────────────────────────
DB_CONFIG = {
    "host":     "localhost",
    "database": "sales_db",
    "user":     "your_user",
    "password": "your_password",
    "port":     5432,
}
OUTPUT_DIR   = "output"
HEADER_COLOR = "1F4E79"
ALT_ROW_COLOR = "EBF2FA"


# ── Helpers ───────────────────────────────────────────────
def run_query(sql_file: str) -> pd.DataFrame:
    """Execute a .sql file and return results as a DataFrame."""
    with open(sql_file) as f:
        query = f.read()
    conn = psycopg2.connect(**DB_CONFIG)
    try:
        df = pd.read_sql(query, conn)
    finally:
        conn.close()
    return df


def style_header_row(ws, row_num: int, col_count: int):
    header_font  = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill  = PatternFill("solid", fgColor=HEADER_COLOR)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, col_count + 1):
        cell            = ws.cell(row=row_num, column=col)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center_align
        cell.border     = border


def style_data_rows(ws, start_row: int, end_row: int, col_count: int):
    alt_fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if r % 2 == 0:
                cell.fill = alt_fill


def auto_column_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)


def write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, title: str):
    ws = wb.create_sheet(title=sheet_name)

    # Report title
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color=HEADER_COLOR, name="Arial")

    ws["A2"] = f"Generated on: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="888888", name="Arial")

    # Blank spacer row
    ws.append([])

    # Write DataFrame (header + data)
    header_row = 4
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=header_row):
        ws.append(row)
        if r_idx == header_row:
            style_header_row(ws, r_idx, len(df.columns))

    # Style data rows
    style_data_rows(ws, header_row + 1, header_row + len(df), len(df.columns))

    # Row height for header
    ws.row_dimensions[header_row].height = 28

    auto_column_width(ws)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


# ── Main ──────────────────────────────────────────────────
def generate_report() -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    output_path = os.path.join(OUTPUT_DIR, f"MIS_Report_{today}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("Weekly Summary",   "queries/weekly_summary.sql",   "Weekly KPI Summary"),
        ("Region Breakdown", "queries/region_breakdown.sql", "Region-wise Performance"),
        ("Top Products",     "queries/top_products.sql",     "Top 10 Products by Revenue"),
    ]

    for sheet_name, query_file, title in sheets:
        print(f"  Running query: {query_file}")
        df = run_query(query_file)
        write_sheet(wb, sheet_name, df, title)
        print(f"  ✓ {len(df):,} rows → sheet '{sheet_name}'")

    wb.save(output_path)
    print(f"\n✅ Report saved: {output_path}")
    return output_path


if __name__ == "__main__":
    print("=== MIS Report Generator ===")
    generate_report()
